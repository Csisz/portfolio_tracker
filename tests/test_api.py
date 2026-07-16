"""
Flask API végpont tesztek – session auth + mock-ok.
"""
import sys, os
from io import BytesIO
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import pytest
from unittest.mock import patch

import services.db as db_module
from services.settings_store import invalidate_cache, save_setting
import app as flask_app


@pytest.fixture(autouse=True)
def _reset_engine_before_each(monkeypatch):
    """Minden teszt előtt törli az engine cache-t."""
    monkeypatch.delenv("DATABASE_URL", raising=False)
    monkeypatch.delenv("APP_ENV", raising=False)
    db_module.reset_engine()
    yield
    db_module.reset_engine()


@pytest.fixture()
def client(tmp_path, monkeypatch):
    """Test client: ideiglenes DB, bejelentkezve admin/admin."""
    db_path = str(tmp_path / "test_api.db")
    monkeypatch.setattr(db_module, "DB_PATH", db_path)
    monkeypatch.setattr(db_module, "_PORTFOLIO_JSON", str(tmp_path / "portfolio.json"))
    monkeypatch.setattr(db_module, "_SYMBOLS_JSON", str(tmp_path / "symbols.json"))
    db_module.reset_engine()  # engine cache törlése – minden teszt saját DB-t kap

    # DB + admin user inicializálás
    db_module.init_db("admin", "adminpass")

    flask_app.app.config["TESTING"] = True
    flask_app.app.config["SECRET_KEY"] = "test-secret-abc"

    with flask_app.app.test_client() as c:
        # Login
        c.post("/login", data={"username": "admin", "password": "adminpass"})
        yield c


@pytest.fixture()
def unauth_client(tmp_path, monkeypatch):
    """Test client: nincs bejelentkezve."""
    db_path = str(tmp_path / "test_unauth.db")
    monkeypatch.setattr(db_module, "DB_PATH", db_path)
    monkeypatch.setattr(db_module, "_PORTFOLIO_JSON", str(tmp_path / "p.json"))
    monkeypatch.setattr(db_module, "_SYMBOLS_JSON", str(tmp_path / "s.json"))
    db_module.reset_engine()
    db_module.init_db("admin", "adminpass")
    flask_app.app.config["TESTING"] = True
    flask_app.app.config["SECRET_KEY"] = "test-secret-abc"
    with flask_app.app.test_client() as c:
        yield c


# ===========================================================================
# Auth
# ===========================================================================

def test_login_redirects_to_index(tmp_path, monkeypatch):
    db_path = str(tmp_path / "login.db")
    monkeypatch.setattr(db_module, "DB_PATH", db_path)
    monkeypatch.setattr(db_module, "_PORTFOLIO_JSON", str(tmp_path / "p.json"))
    monkeypatch.setattr(db_module, "_SYMBOLS_JSON", str(tmp_path / "s.json"))
    db_module.init_db("testuser", "testpass")
    flask_app.app.config["TESTING"] = True
    flask_app.app.config["SECRET_KEY"] = "test"
    with flask_app.app.test_client() as c:
        r = c.post("/login", data={"username": "testuser", "password": "testpass"})
        assert r.status_code in (200, 302)


def test_login_wrong_password_stays_on_login(tmp_path, monkeypatch):
    db_path = str(tmp_path / "loginwrong.db")
    monkeypatch.setattr(db_module, "DB_PATH", db_path)
    monkeypatch.setattr(db_module, "_PORTFOLIO_JSON", str(tmp_path / "p.json"))
    monkeypatch.setattr(db_module, "_SYMBOLS_JSON", str(tmp_path / "s.json"))
    db_module.init_db("testuser", "correct")
    flask_app.app.config["TESTING"] = True
    flask_app.app.config["SECRET_KEY"] = "test"
    with flask_app.app.test_client() as c:
        r = c.post("/login", data={"username": "testuser", "password": "wrong"})
        assert r.status_code == 200
        assert b"Hib" in r.data  # "Hibás jelszó"


def test_protected_route_requires_login(unauth_client):
    r = unauth_client.get("/api/fx")
    assert r.status_code == 401
    d = r.get_json()
    assert d.get("login_required") is True


def test_protected_portfolio_requires_login(unauth_client):
    r = unauth_client.get("/api/portfolio")
    assert r.status_code == 401


# ===========================================================================
# /api/fx
# ===========================================================================

def _mnb_ok():
    return {
        "fx": {"EUR/HUF": 395.40, "USD/HUF": 362.80, "USD/EUR": 0.9178},
        "errors": [],
        "source": "MNB",
        "timestamp": "2026-06-04T10:00:00",
        "date": "2026-06-04",
    }


def test_fx_structure(client):
    with patch("app.get_fx_rates", return_value=_mnb_ok()):
        r = client.get("/api/fx")
    assert r.status_code == 200
    d = r.get_json()
    assert "fx" in d
    assert "errors" in d
    assert "source" in d
    assert "timestamp" in d


def test_fx_response_is_frontend_compatible_with_market_and_official(client):
    payload = {
        "mode": "market",
        "requested_mode": "market",
        "fx": {"EUR/HUF": 353.9, "USD/HUF": 304.21},
        "market": {"EUR/HUF": 353.9, "USD/HUF": 304.21, "source": "Yahoo Finance FX", "timestamp": "2026-06-04T16:15:00"},
        "official": {"EUR/HUF": 355.14, "USD/HUF": 305.81, "source": "MNB", "date": "2026-06-04", "timestamp": "2026-06-04T10:00:00"},
        "errors": [],
        "source": "Yahoo Finance FX",
        "timestamp": "2026-06-04T16:15:00",
    }
    with patch("app.get_fx_rates", return_value=payload):
        r = client.get("/api/fx")
    assert r.status_code == 200
    d = r.get_json()
    assert d["fx"]["EUR/HUF"] == 353.9
    assert d["market"]["USD/HUF"] == 304.21
    assert d["official"]["EUR/HUF"] == 355.14
    assert d["source"] == "Yahoo Finance FX"


def test_fx_has_eur_usd(client):
    with patch("app.get_fx_rates", return_value=_mnb_ok()):
        r = client.get("/api/fx")
    d = r.get_json()
    assert "EUR/HUF" in d["fx"]
    assert "USD/HUF" in d["fx"]


def test_fx_no_500_when_mnb_down(client):
    with patch("app.get_fx_rates", return_value={
        "fx": {}, "errors": ["MNB nem elérhető"], "source": "none",
        "timestamp": "2026-06-04T10:00:00"
    }):
        r = client.get("/api/fx")
    assert r.status_code == 200


# ===========================================================================
# /api/search
# ===========================================================================

def test_search_otp_never_500(client):
    with patch("services.symbol_resolver._search_yahoo", return_value=([], ["rate limit"])):
        r = client.get("/api/search/OTP")
    assert r.status_code == 200


def test_search_otp_has_results(client):
    with patch("services.symbol_resolver._search_yahoo", return_value=([], ["rate limit"])):
        r = client.get("/api/search/OTP")
    d = r.get_json()
    assert "results" in d
    tickers = [x["ticker"] for x in d["results"]]
    assert "OTP.BD" in tickers


def test_search_apple_fallback(client):
    with patch("services.symbol_resolver._search_yahoo", return_value=([], ["rate limit"])):
        r = client.get("/api/search/Apple")
    d = r.get_json()
    tickers = [x["ticker"] for x in d["results"]]
    assert "AAPL" in tickers


def test_search_bmw_fallback(client):
    with patch("services.symbol_resolver._search_yahoo", return_value=([], ["rate limit"])):
        r = client.get("/api/search/BMW")
    d = r.get_json()
    tickers = [x["ticker"] for x in d["results"]]
    assert "BMW.DE" in tickers


def test_search_unknown_returns_200_not_500(client):
    with patch("services.symbol_resolver._search_yahoo", return_value=([], [])):
        r = client.get("/api/search/XYZUNKNOWN999")
    assert r.status_code == 200
    d = r.get_json()
    assert "results" in d


# ===========================================================================
# /api/prices
# ===========================================================================

def _mock_prices(tickers, force_refresh=False):
    prices = {}
    errors = []
    for t in tickers:
        if t == "AAPL":
            prices[t] = {"price": 195.23, "currency": "USD",
                         "source": "Yahoo Finance", "quote_time": "2026-06-04T10:00:00",
                         "received_at": "2026-06-04T10:00:06", "timestamp": "2026-06-04T10:00:00",
                         "stale": False, "delayed": False, "market_state": "REGULAR"}
        else:
            errors.append({"ticker": t, "message": "Árfolyam most nem elérhető."})
    return {"prices": prices, "errors": errors, "timestamp": "2026-06-04T10:00:00", "source": "Yahoo Finance"}


def test_prices_structure(client):
    with patch("app.get_prices_for_tickers", side_effect=_mock_prices):
        r = client.post("/api/prices", json={"tickers": ["AAPL", "OTP.BD"]})
    assert r.status_code == 200
    d = r.get_json()
    assert "prices" in d
    assert "errors" in d


def test_prices_passes_force_refresh_flag(client):
    with patch("app.get_prices_for_tickers", return_value={
        "prices": {}, "errors": [], "timestamp": "2026-06-04T10:00:00", "source": "none"
    }) as mocked:
        r = client.post("/api/prices", json={"tickers": ["AAPL"], "force_refresh": True})
    assert r.status_code == 200
    mocked.assert_called_once_with(["AAPL"], force_refresh=True)


def test_prices_partial_failure(client):
    with patch("app.get_prices_for_tickers", side_effect=_mock_prices):
        r = client.post("/api/prices", json={"tickers": ["AAPL", "OTP.BD"]})
    d = r.get_json()
    assert "AAPL" in d["prices"]
    assert any(e["ticker"] == "OTP.BD" for e in d["errors"])


def test_prices_empty_tickers_400(client):
    r = client.post("/api/prices", json={"tickers": []})
    assert r.status_code == 400


def test_prices_no_crash_on_error(client):
    with patch("app.get_prices_for_tickers", return_value={
        "prices": {}, "errors": [{"ticker": "X", "message": "hiba"}],
        "timestamp": "...", "source": "none"
    }):
        r = client.post("/api/prices", json={"tickers": ["X"]})
    assert r.status_code == 200


def test_prices_uses_portfolio_last_price_cache_before_error(client):
    client.post("/api/portfolio", json=[{
        "ticker": "OTP.BD",
        "name": "OTP Bank",
        "qty": 1,
        "currency": "HUF",
    }])
    uid = db_module.get_user_by_username("admin")["id"]
    db_module.update_item_last_price(uid, "OTP.BD", 40850.0, "HUF", "Stooq", "2026-06-05T10:00:00")

    with patch("app.get_prices_for_tickers", return_value={
        "prices": {},
        "errors": [{"ticker": "OTP.BD", "message": "Árfolyam most nem elérhető."}],
        "timestamp": "2026-06-05T10:01:00",
        "source": "none",
    }):
        r = client.post("/api/prices", json={"tickers": ["OTP.BD"]})

    assert r.status_code == 200
    d = r.get_json()
    assert d["errors"] == []
    assert d["prices"]["OTP.BD"]["price"] == 40850.0
    assert d["prices"]["OTP.BD"]["stale"] is True
    assert d["prices"]["OTP.BD"]["delayed"] is True
    assert d["prices"]["OTP.BD"]["source"] == "Utolsó ismert árfolyam"
    assert d["prices"]["OTP.BD"]["quote_time"] == "2026-06-05T10:00:00"
    assert d["prices"]["OTP.BD"]["timestamp"] == "2026-06-05T10:00:00"


# ===========================================================================
# /api/portfolio
# ===========================================================================

def test_portfolio_get_empty(client):
    r = client.get("/api/portfolio")
    assert r.status_code == 200
    d = r.get_json()
    assert isinstance(d, list)


def test_portfolio_save_and_retrieve(client):
    items = [{"ticker": "AAPL", "name": "Apple", "qty": 5}]
    r = client.post("/api/portfolio", json=items)
    assert r.status_code == 200
    assert r.get_json()["ok"] is True

    r2 = client.get("/api/portfolio")
    d = r2.get_json()
    assert len(d) == 1
    assert d[0]["ticker"] == "AAPL"


def test_portfolio_delete_by_id(client):
    # Hozzáadás
    client.post("/api/portfolio", json=[{"ticker": "MSFT", "name": "Microsoft", "qty": 3}])
    portfolio = client.get("/api/portfolio").get_json()
    assert len(portfolio) == 1
    item_id = portfolio[0]["id"]

    r = client.delete(f"/api/portfolio/{item_id}")
    assert r.status_code == 200
    assert r.get_json()["ok"] is True

    portfolio_after = client.get("/api/portfolio").get_json()
    assert len(portfolio_after) == 0


def test_portfolio_save_purchase_fields(client):
    items = [{
        "ticker": "AAPL",
        "name": "Apple",
        "qty": 5,
        "purchase_price": 150,
        "purchase_date": "2024-01-15",
        "purchase_cost": 7.5,
        "purchase_price_source": "manual",
    }]
    r = client.post("/api/portfolio", json=items)
    assert r.status_code == 200
    saved = client.get("/api/portfolio").get_json()[0]
    assert saved["purchase_price"] == 150.0
    assert saved["purchase_date"] == "2024-01-15"
    assert saved["purchase_cost"] == 7.5
    assert saved["purchase_price_source"] == "manual"


def test_portfolio_patch_purchase_fields(client):
    client.post("/api/portfolio", json=[{"ticker": "AAPL", "name": "Apple", "qty": 3}])
    item_id = client.get("/api/portfolio").get_json()[0]["id"]
    r = client.patch(f"/api/portfolio/{item_id}", json={
        "purchase_price": 175.25,
        "purchase_date": "2024-02-01",
        "purchase_cost": 3.25,
        "purchase_price_source": "historical",
    })
    assert r.status_code == 200
    d = r.get_json()
    assert d["ok"] is True
    assert d["item"]["purchase_price"] == 175.25
    assert d["item"]["purchase_date"] == "2024-02-01"
    assert d["item"]["purchase_cost"] == 3.25
    saved = client.get("/api/portfolio").get_json()[0]
    assert saved["purchase_price_source"] == "historical"


def test_portfolio_patch_quantity_and_purchase_fields_together_keeps_order(client):
    client.post("/api/portfolio", json=[
        {"ticker": "AAPL", "name": "Apple", "qty": 1},
        {"ticker": "MSFT", "name": "Microsoft", "qty": 2},
    ])
    before = client.get("/api/portfolio").get_json()
    item_id = before[0]["id"]
    r = client.patch(f"/api/portfolio/{item_id}", json={
        "qty": 3, "purchase_price": 175.5, "purchase_date": "2024-02-03", "purchase_cost": 4.5
    })
    assert r.status_code == 200
    after = client.get("/api/portfolio").get_json()
    assert [item["id"] for item in after] == [item["id"] for item in before]
    assert after[0]["qty"] == 3.0
    assert after[0]["purchase_price"] == 175.5
    assert after[0]["purchase_date"] == "2024-02-03"
    assert after[0]["purchase_cost"] == 4.5
    assert after[0]["purchase_price_source"] == "manual"


def test_portfolio_patch_validates_and_can_clear_optional_fields(client):
    client.post("/api/portfolio", json=[{
        "ticker": "AAPL", "name": "Apple", "qty": 1,
        "purchase_price": 150, "purchase_date": "2024-01-01", "purchase_cost": 2,
    }])
    item_id = client.get("/api/portfolio").get_json()[0]["id"]
    for payload in ({"qty": 0}, {"purchase_price": -1}, {"purchase_cost": -1}, {"purchase_date": "01-02-2024"}):
        assert client.patch(f"/api/portfolio/{item_id}", json=payload).status_code == 400
    r = client.patch(f"/api/portfolio/{item_id}", json={
        "purchase_price": None, "purchase_date": "", "purchase_cost": None, "not_a_column": "ignored"
    })
    assert r.status_code == 200
    saved = r.get_json()["item"]
    assert saved["purchase_price"] is None
    assert saved["purchase_date"] is None
    assert saved["purchase_cost"] is None


def test_cash_can_be_added_as_separate_lots_and_quoted_without_external_calls(client):
    with patch("services.stocks._fetch_price_yfinance") as yahoo, patch("services.stocks._fetch_price_stooq") as stooq:
        first = client.post("/api/cash", json={"currency": "HUF", "amount": 100000})
        second = client.post("/api/cash", json={"currency": "HUF", "amount": 25000})
        quote = client.post("/api/prices", json={"tickers": ["CASH-HUF"]})
    assert first.status_code == second.status_code == quote.status_code == 200
    yahoo.assert_not_called()
    stooq.assert_not_called()
    portfolio = client.get("/api/portfolio").get_json()
    assert len(portfolio) == 2
    assert [item["qty"] for item in portfolio] == [100000.0, 25000.0]
    assert all(item["purchase_price"] == 1.0 for item in portfolio)
    assert quote.get_json()["prices"]["CASH-HUF"]["price"] == 1.0


def test_huf_cash_total_needs_no_fx(client):
    from app import _portfolio_total_huf
    total = _portfolio_total_huf(
        [{"ticker": "CASH-HUF", "qty": 100000, "currency": "HUF"}],
        {"CASH-HUF": {"price": 1, "currency": "HUF"}},
        {},
    )
    assert total == 100000.0


def test_portfolio_rejects_negative_purchase_cost(client):
    r = client.post("/api/portfolio", json=[{
        "ticker": "AAPL",
        "name": "Apple",
        "qty": 1,
        "purchase_price": 150,
        "purchase_cost": -1,
    }])
    assert r.status_code == 400

    client.post("/api/portfolio", json=[{"ticker": "AAPL", "name": "Apple", "qty": 1}])
    item_id = client.get("/api/portfolio").get_json()[0]["id"]
    r = client.patch(f"/api/portfolio/{item_id}", json={"purchase_cost": -1})
    assert r.status_code == 400


def test_add_manual_accepts_purchase_price(client):
    with patch("app.get_ticker_info", return_value={
        "ticker": "AAPL",
        "name": "Apple Inc.",
        "currency": "USD",
        "exchange": "NASDAQ",
        "last_price": 200.0,
    }):
        r = client.post("/api/add_manual", json={
            "ticker": "AAPL",
            "qty": 2,
            "purchase_price": 180,
            "purchase_date": "2024-01-15",
            "purchase_cost": 1.25,
            "purchase_price_source": "manual",
        })
    assert r.status_code == 200
    saved = client.get("/api/portfolio").get_json()[0]
    assert saved["purchase_price"] == 180.0
    assert saved["purchase_date"] == "2024-01-15"
    assert saved["purchase_cost"] == 1.25


def test_add_manual_same_ticker_creates_separate_lots(client):
    with patch("app.get_ticker_info", return_value=None):
        r1 = client.post("/api/add_manual", json={
            "ticker": "OTP.BD",
            "qty": 1,
            "purchase_price": 40250,
        })
        r2 = client.post("/api/add_manual", json={
            "ticker": "OTP.BD",
            "qty": 1,
            "purchase_price": 40600,
        })

    assert r1.status_code == 200
    assert r2.status_code == 200
    d2 = r2.get_json()
    assert "merged" not in d2
    assert d2["message"] == "Új vételi tétel hozzáadva."

    portfolio = client.get("/api/portfolio").get_json()
    assert len(portfolio) == 2
    assert [item["ticker"] for item in portfolio] == ["OTP.BD", "OTP.BD"]
    assert [item["qty"] for item in portfolio] == [1.0, 1.0]
    assert [item["purchase_price"] for item in portfolio] == [40250.0, 40600.0]


def test_add_manual_mol_same_ticker_creates_separate_lots(client):
    with patch("app.get_ticker_info", return_value=None):
        client.post("/api/add_manual", json={
            "ticker": "MOL.BD",
            "qty": 2,
            "purchase_price": 3000,
        })
        r = client.post("/api/add_manual", json={
            "ticker": "MOL",
            "qty": 1,
            "purchase_price": 3300,
        })

    assert r.status_code == 200
    portfolio = client.get("/api/portfolio").get_json()
    assert len(portfolio) == 2
    assert [item["ticker"] for item in portfolio] == ["MOL.BD", "MOL.BD"]
    assert [item["qty"] for item in portfolio] == [2.0, 1.0]
    assert [item["purchase_price"] for item in portfolio] == [3000.0, 3300.0]


def test_add_manual_new_ticker_still_creates_new_row(client):
    with patch("app.get_ticker_info", return_value=None):
        client.post("/api/add_manual", json={
            "ticker": "OTP.BD",
            "qty": 1,
            "purchase_price": 40250,
        })
        client.post("/api/add_manual", json={
            "ticker": "MOL.BD",
            "qty": 1,
            "purchase_price": 3000,
        })

    portfolio = client.get("/api/portfolio").get_json()
    assert {item["ticker"] for item in portfolio} == {"OTP.BD", "MOL.BD"}


def test_new_manual_lots_append_even_with_older_purchase_date(client):
    with patch("app.get_ticker_info", return_value=None):
        client.post("/api/add_manual", json={
            "ticker": "AAPL",
            "qty": 1,
            "purchase_price": 190,
            "purchase_date": "2024-03-01",
        })
        client.post("/api/add_manual", json={
            "ticker": "AAPL",
            "qty": 1,
            "purchase_price": 170,
            "purchase_date": "2024-01-15",
        })

    portfolio = client.get("/api/portfolio").get_json()
    assert [item["purchase_date"] for item in portfolio] == ["2024-03-01", "2024-01-15"]
    assert [item["purchase_price"] for item in portfolio] == [190.0, 170.0]
    assert [item["display_order"] for item in portfolio] == [0, 1]


def test_portfolio_reorder_saves_manual_order(client):
    client.post("/api/portfolio", json=[
        {"ticker": "AAPL", "name": "Apple", "qty": 1},
        {"ticker": "MSFT", "name": "Microsoft", "qty": 1},
        {"ticker": "TSLA", "name": "Tesla", "qty": 1},
    ])
    portfolio = client.get("/api/portfolio").get_json()
    ordered_ids = [portfolio[2]["id"], portfolio[0]["id"], portfolio[1]["id"]]

    r = client.patch("/api/portfolio/reorder", json={"ordered_ids": ordered_ids})

    assert r.status_code == 200
    data = r.get_json()
    assert data["ok"] is True
    assert [item["id"] for item in data["portfolio"]] == ordered_ids
    assert [item["ticker"] for item in client.get("/api/portfolio").get_json()] == ["TSLA", "AAPL", "MSFT"]


def test_portfolio_reorder_duplicate_ticker_lots_by_id(client):
    with patch("app.get_ticker_info", return_value=None):
        client.post("/api/add_manual", json={"ticker": "OTP.BD", "qty": 1, "purchase_price": 40250})
        client.post("/api/add_manual", json={"ticker": "OTP.BD", "qty": 2, "purchase_price": 40600})
    portfolio = client.get("/api/portfolio").get_json()
    ordered_ids = [portfolio[1]["id"], portfolio[0]["id"]]

    r = client.patch("/api/portfolio/reorder", json={"ordered_ids": ordered_ids})

    assert r.status_code == 200
    updated = client.get("/api/portfolio").get_json()
    assert [item["id"] for item in updated] == ordered_ids
    assert [item["qty"] for item in updated] == [2.0, 1.0]


def test_user_cannot_reorder_another_users_items(client):
    client.post("/api/portfolio", json=[{"ticker": "AAPL", "name": "Apple", "qty": 1}])
    admin_item = client.get("/api/portfolio").get_json()[0]
    db_module.create_user("other", "otherpass")
    client.get("/logout")
    client.post("/login", data={"username": "other", "password": "otherpass"})
    client.post("/api/portfolio", json=[{"ticker": "MSFT", "name": "Microsoft", "qty": 1}])
    other_item = client.get("/api/portfolio").get_json()[0]

    r = client.patch("/api/portfolio/reorder", json={"ordered_ids": [other_item["id"], admin_item["id"]]})

    assert r.status_code == 403
    assert [item["ticker"] for item in client.get("/api/portfolio").get_json()] == ["MSFT"]


def test_add_manual_existing_without_purchase_price_still_creates_new_lot(client):
    client.post("/api/portfolio", json=[{
        "ticker": "OTP.BD",
        "name": "OTP Bank",
        "qty": 1,
    }])

    with patch("app.get_ticker_info", return_value=None):
        r = client.post("/api/add_manual", json={
            "ticker": "OTP.BD",
            "qty": 1,
            "purchase_price": 40600,
        })

    assert r.status_code == 200
    portfolio = client.get("/api/portfolio").get_json()
    assert len(portfolio) == 2
    assert portfolio[0]["qty"] == 1.0
    assert portfolio[0]["purchase_price"] is None
    assert portfolio[1]["qty"] == 1.0
    assert portfolio[1]["purchase_price"] == 40600.0


def test_portfolio_patch_duplicate_ticker_updates_only_selected_lot(client):
    with patch("app.get_ticker_info", return_value=None):
        client.post("/api/add_manual", json={
            "ticker": "OTP.BD",
            "qty": 1,
            "purchase_price": 40250,
            "purchase_date": "2024-02-01",
        })
        client.post("/api/add_manual", json={
            "ticker": "OTP.BD",
            "qty": 1,
            "purchase_price": 40600,
            "purchase_date": "2024-02-10",
        })
    portfolio = client.get("/api/portfolio").get_json()
    second_id = portfolio[1]["id"]

    r_qty = client.put(f"/api/portfolio/{second_id}", json={"qty": 3})
    r_price = client.patch(f"/api/portfolio/{second_id}", json={
        "purchase_price": 40700,
        "purchase_date": "2024-02-15",
        "purchase_cost": 125,
    })

    assert r_qty.status_code == 200
    assert r_price.status_code == 200
    updated = client.get("/api/portfolio").get_json()
    assert updated[0]["qty"] == 1.0
    assert updated[0]["purchase_price"] == 40250.0
    assert updated[0]["purchase_date"] == "2024-02-01"
    assert updated[0]["purchase_cost"] is None
    assert updated[1]["qty"] == 3.0
    assert updated[1]["purchase_price"] == 40700.0
    assert updated[1]["purchase_date"] == "2024-02-15"
    assert updated[1]["purchase_cost"] == 125.0


def test_portfolio_delete_duplicate_ticker_removes_only_selected_lot(client):
    with patch("app.get_ticker_info", return_value=None):
        client.post("/api/add_manual", json={
            "ticker": "OTP.BD",
            "qty": 1,
            "purchase_price": 40250,
        })
        client.post("/api/add_manual", json={
            "ticker": "OTP.BD",
            "qty": 1,
            "purchase_price": 40600,
        })
    portfolio = client.get("/api/portfolio").get_json()
    first_id = portfolio[0]["id"]

    r = client.delete(f"/api/portfolio/{first_id}")

    assert r.status_code == 200
    remaining = client.get("/api/portfolio").get_json()
    assert len(remaining) == 1
    assert remaining[0]["ticker"] == "OTP.BD"
    assert remaining[0]["purchase_price"] == 40600.0


def test_add_manual_defaults_purchase_price_to_current_price(client):
    with patch("app.get_ticker_info", return_value=None), \
         patch("app.get_prices_for_tickers", return_value={
             "prices": {"OTP.BD": {
                 "price": 40850.0,
                 "currency": "HUF",
                 "source": "Stooq",
                 "timestamp": "2026-06-05T10:00:00",
             }},
             "errors": [],
             "timestamp": "2026-06-05T10:00:00",
             "source": "Stooq",
         }):
        r = client.post("/api/add_manual", json={
            "ticker": "OTP.BD",
            "qty": 1,
        })

    assert r.status_code == 200
    saved = client.get("/api/portfolio").get_json()[0]
    assert saved["purchase_price"] == 40850.0


def test_add_manual_normalizes_otp_and_uses_price_service(client):
    with patch("app.get_ticker_info", return_value=None), \
         patch("app.get_prices_for_tickers", return_value={
             "prices": {"OTP.BD": {
                 "price": 40850.0,
                 "currency": "HUF",
                 "source": "Stooq",
                 "timestamp": "2026-06-05T10:00:00",
             }},
             "errors": [],
             "timestamp": "2026-06-05T10:00:00",
             "source": "Stooq",
         }) as mock_prices:
        r = client.post("/api/add_manual", json={
            "ticker": "OTP",
            "qty": 1,
        })

    assert r.status_code == 200
    mock_prices.assert_called_once_with(["OTP.BD"])
    saved = client.get("/api/portfolio").get_json()[0]
    assert saved["ticker"] == "OTP.BD"
    assert saved["purchase_price"] == 40850.0


def test_add_manual_rejects_missing_purchase_price_when_no_price_fallback(client):
    with patch("app.get_ticker_info", return_value=None), \
         patch("app.get_prices_for_tickers", return_value={
             "prices": {},
             "errors": [{"ticker": "MOL.BD", "message": "Árfolyam most nem elérhető."}],
             "timestamp": "2026-06-05T10:00:00",
             "source": "none",
         }):
        r = client.post("/api/add_manual", json={
            "ticker": "MOL.BD",
            "qty": 3,
        })

    assert r.status_code == 400
    data = r.get_json()
    assert data["ok"] is False
    assert data["error"] == "Vételi ár megadása kötelező."
    assert client.get("/api/portfolio").get_json() == []


def test_add_manual_rejects_non_positive_purchase_price(client):
    for value in ("", 0, -1, "abc"):
        with patch("app.get_ticker_info", return_value=None), \
             patch("app.get_prices_for_tickers", return_value={
                 "prices": {},
                 "errors": [],
                 "timestamp": "2026-06-05T10:00:00",
                 "source": "none",
             }):
            r = client.post("/api/add_manual", json={
                "ticker": "MOL.BD",
                "qty": 3,
                "purchase_price": value,
            })

        assert r.status_code == 400
        assert r.get_json()["error"] == "Vételi ár megadása kötelező."

    assert client.get("/api/portfolio").get_json() == []


def test_add_manual_uses_cached_price_from_price_service(client):
    with patch("app.get_ticker_info", return_value=None), \
         patch("app.get_prices_for_tickers", return_value={
             "prices": {"MOL.BD": {
                 "price": 3100.0,
                 "currency": "HUF",
                 "source": "stale",
                 "timestamp": "2026-06-05T10:00:00",
                 "stale": True,
             }},
             "errors": [],
             "timestamp": "2026-06-05T10:00:00",
             "source": "Yahoo Finance/cache",
         }):
        r = client.post("/api/add_manual", json={
            "ticker": "MOL.BD",
            "qty": 3,
        })

    assert r.status_code == 200
    data = r.get_json()
    assert "warning" not in data
    saved = client.get("/api/portfolio").get_json()[0]
    assert saved["purchase_price"] == 3100.0


def test_add_manual_keeps_metadata_from_ticker_info_with_price_service(client):
    with patch("app.get_ticker_info", return_value={
        "ticker": "OTP.BD",
        "name": "OTP Bank",
        "currency": "HUF",
        "exchange": "BUD",
    }), patch("app.get_prices_for_tickers", return_value={
        "prices": {"OTP.BD": {
            "price": 40850.0,
            "currency": "HUF",
            "source": "Stooq",
            "timestamp": "2026-06-05T10:00:00",
        }},
        "errors": [],
        "timestamp": "2026-06-05T10:00:00",
        "source": "Stooq",
    }):
        r = client.post("/api/add_manual", json={
            "ticker": "OTP.BD",
            "qty": 1,
        })

    assert r.status_code == 200
    saved = client.get("/api/portfolio").get_json()[0]
    assert saved["name"] == "OTP Bank"
    assert saved["purchase_price"] == 40850.0


def test_price_history_endpoint(client):
    with patch("app.get_historical_price", return_value={
        "ok": True,
        "ticker": "AAPL",
        "requested_date": "2024-01-15",
        "used_date": "2024-01-12",
        "price": 185.5,
        "currency": "USD",
        "source": "Yahoo Finance",
    }) as mock_history:
        r = client.get("/api/price-history?ticker=AAPL&date=2024-01-15")
    assert r.status_code == 200
    d = r.get_json()
    assert d["price"] == 185.5
    assert d["used_date"] == "2024-01-12"
    mock_history.assert_called_once_with("AAPL", "2024-01-15")


# ===========================================================================
# /api/export/xlsx – MIME type
# ===========================================================================

def test_export_xlsx_mime_type(client):
    # Portfolio feltöltés
    client.post("/api/portfolio", json=[{"ticker": "AAPL", "name": "Apple", "qty": 3}])
    with patch("app.get_prices_for_tickers", return_value={
        "prices": {"AAPL": {"price": 195.0, "currency": "USD", "source": "Yahoo Finance", "timestamp": "..."}},
        "errors": [], "timestamp": "...", "source": "Yahoo Finance"
    }), patch("app.get_fx_rates", return_value=_mnb_ok()):
        r = client.get("/api/export/xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.content_type


def test_export_xlsx_includes_purchase_fields_and_cost_adjusted_profit(client):
    client.post("/api/portfolio", json=[{
        "ticker": "AAPL",
        "name": "Apple",
        "qty": 3,
        "purchase_price": 180,
        "purchase_date": "2024-01-15",
        "purchase_cost": 5,
    }])
    with patch("app.get_prices_for_tickers", return_value={
        "prices": {"AAPL": {"price": 200.0, "currency": "USD", "source": "Yahoo Finance", "timestamp": "..."}},
        "errors": [], "timestamp": "...", "source": "Yahoo Finance"
    }), patch("app.get_fx_rates", return_value=_mnb_ok()):
        r = client.get("/api/export/xlsx")

    assert r.status_code == 200
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(r.data))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert "Vétel dátuma" in headers
    assert "Vételi költség" in headers
    assert "Nyereség / veszteség" in headers
    row = [cell.value for cell in ws[2]]
    by_header = dict(zip(headers, row))
    assert by_header["Vétel dátuma"] == "2024-01-15"
    assert by_header["Vételi költség"] == 5
    assert by_header["Befektetett érték"] == 545
    assert by_header["Nyereség / veszteség"] == 55
    assert by_header["Hozam %"] == 10.09


def test_export_xlsx_follows_manual_portfolio_order(client):
    client.post("/api/portfolio", json=[
        {"ticker": "AAPL", "name": "Apple", "qty": 1},
        {"ticker": "MSFT", "name": "Microsoft", "qty": 1},
        {"ticker": "TSLA", "name": "Tesla", "qty": 1},
    ])
    portfolio = client.get("/api/portfolio").get_json()
    ordered_ids = [portfolio[1]["id"], portfolio[2]["id"], portfolio[0]["id"]]
    client.patch("/api/portfolio/reorder", json={"ordered_ids": ordered_ids})
    with patch("app.get_prices_for_tickers", return_value={
        "prices": {
            "AAPL": {"price": 200.0, "currency": "USD", "source": "Yahoo Finance", "timestamp": "..."},
            "MSFT": {"price": 300.0, "currency": "USD", "source": "Yahoo Finance", "timestamp": "..."},
            "TSLA": {"price": 400.0, "currency": "USD", "source": "Yahoo Finance", "timestamp": "..."},
        },
        "errors": [], "timestamp": "...", "source": "Yahoo Finance"
    }), patch("app.get_fx_rates", return_value=_mnb_ok()):
        r = client.get("/api/export/xlsx")

    assert r.status_code == 200
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(r.data))
    ws = wb.active
    assert [ws.cell(row=i, column=2).value for i in range(2, 5)] == ["MSFT", "TSLA", "AAPL"]


def test_export_xlsx_empty_portfolio_400(client):
    r = client.get("/api/export/xlsx")
    assert r.status_code == 400


# ===========================================================================
# /api/alerts
# ===========================================================================

def test_alert_check_disabled_does_not_send_email(client):
    save_setting("alerts_enabled", "false")
    invalidate_cache()
    with patch("app.send_email") as mock_send:
        r = client.post("/api/alerts/check")
    assert r.status_code == 200
    d = r.get_json()
    assert d["ok"] is True
    assert d["alerts_enabled"] is False
    assert d["message"] == "Email alerts are disabled"
    mock_send.assert_not_called()


def test_alert_create_disabled_is_blocked(client):
    save_setting("alerts_enabled", "false")
    invalidate_cache()
    r = client.post("/api/alerts", json={
        "alert_type": "portfolio_value_above",
        "threshold": 100000,
        "currency": "HUF",
        "email_to": "test@example.com",
    })
    assert r.status_code == 403
    assert r.get_json()["alerts_enabled"] is False


def test_alert_check_enabled_can_send_email(client):
    save_setting("alerts_enabled", "true")
    invalidate_cache()
    alert = {
        "id": 123,
        "alert_type": "portfolio_value_above",
        "ticker": None,
        "threshold": 100,
        "percent": None,
        "currency": "HUF",
        "email_to": "test@example.com",
        "cooldown_minutes": 60,
        "last_value": None,
        "last_triggered_at": None,
    }
    with patch("app.get_alerts", return_value=[alert]), \
         patch("app.get_portfolio", return_value=[{"ticker": "AAPL", "qty": 1, "currency": "HUF"}]), \
         patch("app.get_prices_for_tickers", return_value={"prices": {"AAPL": {"price": 200, "currency": "HUF"}}}), \
         patch("app.get_fx_rates", return_value={"fx": {}}), \
         patch("app.update_alert_state", return_value=True), \
         patch("app.send_email", return_value=(True, "Email elküldve.")) as mock_send:
        r = client.post("/api/alerts/check")
    assert r.status_code == 200
    d = r.get_json()
    assert d["alerts_enabled"] is True
    assert len(d["triggered"]) == 1
    mock_send.assert_called_once()
