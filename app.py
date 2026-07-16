"""
Zoltán Portfólió Követő
=======================
Indítás: python app.py
Majd böngészőben: http://localhost:5000

Belépési adatok (fejlesztésben): admin / admin
Éles használathoz állítsd be a .env fájlban:
  SECRET_KEY=...
  PORTFOLIO_USERNAME=...
  PORTFOLIO_PASSWORD=...
  DATABASE_URL=postgresql://...
  APP_ENV=production
"""

import functools
import io
import logging
import os
import secrets
import sys
from datetime import datetime, timedelta

from flask import (Flask, Response, flash, jsonify, redirect, render_template,
                   request, send_file, session, url_for)

import services.db as db
from services import symbol_resolver
from services.db import (create_alert, create_user, delete_alert,
                          delete_portfolio_item_by_id, get_alerts,
                          get_all_users, get_audit_logs, get_portfolio,
                          get_stats, get_user_by_id, init_db, insert_portfolio_item,
                          log_event,
                          reorder_portfolio_items,
                          save_full_portfolio, set_alert_active,
                          set_user_password, update_alert_state,
                          update_portfolio_item_by_id,
                          upsert_symbol_cache,
                          update_item_last_price, update_last_login,
                          update_user, verify_password)
from services.fx import get_fx_rates
from services.settings_store import (get_all_settings_with_defaults,
                                      get_bool, get_int, get_setting,
                                      get_setting_bool, get_setting_int,
                                      init_default_settings, save_setting)
from services.stocks import (cash_currency_from_ticker, get_historical_price,
                             get_prices_for_tickers, get_ticker_info,
                             normalize_ticker)
from services.emailer import send_email, smtp_configured

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# .env betöltés
# ---------------------------------------------------------------------------
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ---------------------------------------------------------------------------
# Flask app
# ---------------------------------------------------------------------------
app = Flask(__name__)

_sk = os.environ.get("SECRET_KEY", "").strip()
_app_env = os.environ.get("APP_ENV", "").lower()

if not _sk:
    if _app_env == "production":
        logger.error("SECRET_KEY nincs beállítva production módban! Az alkalmazás nem indul.")
        sys.exit(1)
    _sk = secrets.token_hex(32)
    logger.warning(
        "SECRET_KEY nincs beállítva. Ideiglenes kulcs aktív – session-ök újraindításkor lejárnak. "
        "Generálj: python -c \"import secrets; print(secrets.token_hex(32))\""
    )

app.secret_key = _sk

# ---------------------------------------------------------------------------
# Context processor – globálisan elérhető változók minden template-ben
# ---------------------------------------------------------------------------

@app.context_processor
def inject_user_context():
    return {
        "current_username": session.get("username", ""),
        "current_role": session.get("role", "user"),
        "is_admin": session.get("role") == "admin",
    }


# Cookie biztonság
if _app_env == "production":
    app.config["SESSION_COOKIE_SECURE"] = True
    app.config["SESSION_COOKIE_HTTPONLY"] = True
    app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
else:
    app.config["SESSION_COOKIE_HTTPONLY"] = True
    app.config["SESSION_COOKIE_SAMESITE"] = "Lax"


# ---------------------------------------------------------------------------
# Auth decoratorok
# ---------------------------------------------------------------------------

def login_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            if request.path.startswith("/api/"):
                return jsonify({"error": "Bejelentkezés szükséges", "login_required": True}), 401
            return redirect(url_for("login_page"))

        # Karbantartási mód: csak admin léphet be
        if get_bool("maintenance_mode") and session.get("role") != "admin":
            if request.path.startswith("/api/"):
                return jsonify({"error": "Karbantartás folyamatban. Próbáld később."}), 503
            return render_template("maintenance.html"), 503

        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            if request.path.startswith("/api/"):
                return jsonify({"error": "Bejelentkezés szükséges"}), 401
            return redirect(url_for("login_page"))
        if session.get("role") != "admin":
            if request.path.startswith("/api/"):
                return jsonify({"error": "Nincs jogosultságod ehhez a művelethez."}), 403
            return render_template("403.html"), 403
        return f(*args, **kwargs)
    return decorated


def current_user_id() -> int:
    return session["user_id"]


def current_username() -> str:
    return session.get("username", "")


def current_role() -> str:
    return session.get("role", "user")


def _client_ip() -> str:
    return request.headers.get("X-Forwarded-For", request.remote_addr or "")


def _positive_float(value):
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if number > 0 else None


def _non_negative_float_or_none(value):
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if number >= 0 else None


# ---------------------------------------------------------------------------
# Login / Logout
# ---------------------------------------------------------------------------

@app.route("/login", methods=["GET", "POST"])
def login_page():
    if "user_id" in session:
        return redirect(url_for("index"))

    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = verify_password(username, password)
        if user:
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["role"] = user.get("role", "user")
            update_last_login(user["id"])
            log_event(user["id"], "login", f"Sikeres belépés: {username}", _client_ip())
            return redirect(url_for("index"))
        else:
            log_event(None, "login_failed", f"Sikertelen belépés: {username}", _client_ip())
            error = "Hibás felhasználónév vagy jelszó, vagy az account inaktív."

    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    uid = session.get("user_id")
    uname = session.get("username", "")
    log_event(uid, "logout", f"Kilépés: {uname}", _client_ip())
    session.clear()
    return redirect(url_for("login_page"))


# ---------------------------------------------------------------------------
# Favicon
# ---------------------------------------------------------------------------

@app.route("/favicon.ico")
def favicon():
    return Response(status=204)


# ---------------------------------------------------------------------------
# Főoldal
# ---------------------------------------------------------------------------

@app.route("/")
@login_required
def index():
    portfolio = get_portfolio(current_user_id())
    return render_template(
        "index.html",
        portfolio=portfolio,
        auto_refresh_seconds=get_setting_int("auto_refresh_seconds", 300),
        alerts_enabled=get_setting_bool("alerts_enabled", False),
        username=current_username(),
        role=current_role(),
    )


# ---------------------------------------------------------------------------
# /api/fx
# ---------------------------------------------------------------------------

@app.route("/api/fx")
@login_required
def get_fx():
    result = get_fx_rates(get_setting("fx_rate_mode", "market"))
    result.pop("_raw", None)
    return jsonify(result)


# ---------------------------------------------------------------------------
# /api/prices
# ---------------------------------------------------------------------------

@app.route("/api/prices", methods=["POST"])
@login_required
def api_prices():
    data = request.get_json(silent=True) or {}
    tickers = data.get("tickers", [])
    force_refresh = data.get("force_refresh") is True
    if not isinstance(tickers, list) or not tickers:
        return jsonify({"error": "Nincs ticker megadva", "prices": {}, "errors": []}), 400

    result = get_prices_for_tickers(tickers, force_refresh=force_refresh)

    uid = current_user_id()
    now = _ts()
    portfolio_by_ticker = {}
    try:
        portfolio_by_ticker = {
            normalize_ticker(item.get("ticker", "")): item
            for item in get_portfolio(uid)
        }
    except Exception:
        portfolio_by_ticker = {}

    prices = result.setdefault("prices", {})
    for requested_ticker in tickers:
        response_ticker = str(requested_ticker or "").strip().upper()
        if response_ticker in prices:
            continue
        cached_item = portfolio_by_ticker.get(normalize_ticker(response_ticker))
        if cached_item and cached_item.get("last_price"):
            prices[response_ticker] = {
                "price": cached_item.get("last_price"),
                "currency": cached_item.get("last_price_currency") or cached_item.get("currency"),
                "source": "Utolsó ismert árfolyam",
                "quote_time": cached_item.get("last_price_time"),
                "received_at": now,
                "timestamp": cached_item.get("last_price_time"),
                "stale": True,
                "delayed": True,
                "market_state": "UNKNOWN",
            }

    if result.get("errors"):
        price_keys = {normalize_ticker(t) for t in prices.keys()}
        result["errors"] = [
            err for err in result.get("errors", [])
            if normalize_ticker(err.get("ticker") if isinstance(err, dict) else err) not in price_keys
        ]

    for ticker, pdata in result.get("prices", {}).items():
        if not pdata.get("stale"):
            try:
                update_item_last_price(uid, ticker, pdata["price"], pdata["currency"], pdata["source"], pdata.get("quote_time") or pdata.get("timestamp"))
            except Exception:
                pass
        if pdata.get("stale"):
            continue
        try:
            existing_symbols = db.search_symbols_db(ticker)
            existing_symbol = next(
                (s for s in existing_symbols if normalize_ticker(s.get("ticker", "")) == normalize_ticker(ticker)),
                {},
            )
            upsert_symbol_cache({
                "ticker": ticker,
                "name": existing_symbol.get("name") or ticker,
                "currency": pdata.get("currency"),
                "exchange": existing_symbol.get("exchange") or "",
                "source": pdata.get("source"),
                "query_aliases": existing_symbol.get("query_aliases") or [ticker.lower()],
                "last_price": pdata.get("price"),
                "last_price_currency": pdata.get("currency"),
                "last_price_time": pdata.get("quote_time") or pdata.get("timestamp"),
            })
        except Exception:
            pass

    return jsonify(result)


@app.route("/api/price-history", methods=["GET"])
@login_required
def api_price_history():
    ticker = request.args.get("ticker", "")
    requested_date = request.args.get("date", "")
    result = get_historical_price(ticker, requested_date)
    return jsonify(result), (200 if result.get("ok") else 400)


# ---------------------------------------------------------------------------
# /api/portfolio
# ---------------------------------------------------------------------------

@app.route("/api/portfolio", methods=["GET"])
@login_required
def api_portfolio_get():
    return jsonify(get_portfolio(current_user_id()))


@app.route("/api/portfolio", methods=["POST"])
@login_required
def api_portfolio_save():
    data = request.get_json(silent=True)
    if not isinstance(data, list):
        return jsonify({"ok": False, "error": "Lista szükséges"}), 400
    for item in data:
        if isinstance(item, dict) and "purchase_cost" in item:
            cost = _non_negative_float_or_none(item.get("purchase_cost"))
            if cost is None and item.get("purchase_cost") not in (None, ""):
                return jsonify({"ok": False, "error": "Ervenytelen veteli koltseg"}), 400
    try:
        save_full_portfolio(current_user_id(), data)
        return jsonify({"ok": True})
    except Exception as e:
        logger.error("Portfolio mentési hiba: %s", e)
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/portfolio/reorder", methods=["PATCH"])
@login_required
def api_portfolio_reorder():
    data = request.get_json(silent=True) or {}
    ordered_ids = data.get("ordered_ids")
    if not isinstance(ordered_ids, list):
        return jsonify({"ok": False, "error": "Ervenytelen sorrend"}), 400
    try:
        portfolio = reorder_portfolio_items(current_user_id(), ordered_ids)
    except PermissionError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 403
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400
    log_event(current_user_id(), "portfolio_reorder", f"rows={len(portfolio)}", _client_ip())
    return jsonify({"ok": True, "portfolio": portfolio})


@app.route("/api/portfolio/<int:item_id>", methods=["DELETE"])
@login_required
def api_portfolio_delete(item_id: int):
    ok = delete_portfolio_item_by_id(current_user_id(), item_id)
    if ok:
        log_event(current_user_id(), "portfolio_delete", f"id={item_id}", _client_ip())
    return jsonify({"ok": ok})


@app.route("/api/portfolio/<int:item_id>", methods=["PUT", "PATCH"])
@login_required
def api_portfolio_update(item_id: int):
    data = request.get_json(silent=True) or {}
    allowed = {"qty", "purchase_price", "purchase_date", "purchase_cost", "purchase_price_source"}
    if not allowed.intersection(data):
        return jsonify({"ok": False, "error": "Nincs frissíthető mező."}), 400

    uid = current_user_id()
    portfolio = get_portfolio(uid)
    item = next((i for i in portfolio if i["id"] == item_id), None)
    if not item:
        return jsonify({"ok": False, "error": "A tétel nem található."}), 404

    updated = dict(item)
    if "qty" in data:
        try:
            qty = float(data.get("qty"))
            if qty <= 0:
                raise ValueError()
        except (ValueError, TypeError):
            return jsonify({"ok": False, "error": "A darab / összeg legyen nullánál nagyobb."}), 400
        updated["qty"] = qty
    if "purchase_price" in data:
        raw_price = data.get("purchase_price")
        if raw_price in (None, ""):
            updated["purchase_price"] = None
            updated["purchase_price_source"] = None
        else:
            try:
                price = float(raw_price)
                if price <= 0:
                    raise ValueError()
            except (ValueError, TypeError):
                return jsonify({"ok": False, "error": "A vételi ár legyen nullánál nagyobb."}), 400
            updated["purchase_price"] = price
            updated["purchase_price_source"] = "manual"
    if "purchase_date" in data:
        purchase_date = str(data.get("purchase_date") or "").strip() or None
        if purchase_date:
            try:
                datetime.strptime(purchase_date, "%Y-%m-%d")
            except ValueError:
                return jsonify({"ok": False, "error": "A vétel dátuma ÉÉÉÉ-HH-NN formátumú legyen."}), 400
        updated["purchase_date"] = purchase_date
    if "purchase_cost" in data:
        cost = _non_negative_float_or_none(data.get("purchase_cost"))
        if cost is None and data.get("purchase_cost") not in (None, ""):
            return jsonify({"ok": False, "error": "A vételi költség nem lehet negatív."}), 400
        updated["purchase_cost"] = cost
    if "purchase_price_source" in data:
        updated["purchase_price_source"] = str(data.get("purchase_price_source") or "").strip().lower() or None

    saved = update_portfolio_item_by_id(uid, item_id, updated)
    if not saved:
        return jsonify({"ok": False, "error": "A tétel nem található."}), 404
    log_event(uid, "portfolio_item_update", f"ticker={item['ticker']} id={item_id}", _client_ip())
    return jsonify({"ok": True, "item": saved})


# ---------------------------------------------------------------------------
# /api/search
# ---------------------------------------------------------------------------

@app.route("/api/search/<path:query>")
@login_required
def api_search(query):
    result = symbol_resolver.search(query.strip())
    return jsonify(result)


# ---------------------------------------------------------------------------
# /api/add_manual
# ---------------------------------------------------------------------------

@app.route("/api/add_manual", methods=["POST"])
@login_required
def api_add_manual():
    data = request.get_json(silent=True) or {}
    ticker = normalize_ticker(data.get("ticker", ""))
    name = str(data.get("name", "")).strip() or ticker
    qty_raw = data.get("qty", 1)

    if not ticker:
        return jsonify({"ok": False, "error": "A ticker mező kötelező."}), 400
    try:
        qty = float(qty_raw)
        if qty <= 0:
            raise ValueError()
    except (ValueError, TypeError):
        return jsonify({"ok": False, "error": "Érvénytelen darabszám."}), 400

    currency = str(data.get("currency", "")).strip().upper() or None
    exchange = str(data.get("exchange", "")).strip()

    cash_currency = cash_currency_from_ticker(ticker)
    if cash_currency:
        currency = cash_currency
        name = f"Készpénz ({cash_currency})"
        exchange = ""
        data["purchase_price"] = 1
        data["purchase_cost"] = 0
        data["purchase_price_source"] = "manual"

    validated = False
    validation_warning = None
    info = None
    try:
        if cash_currency:
            info = None
            validated = True
        else:
            info = get_ticker_info(ticker)
        if info:
            validated = True
            if not currency:
                currency = info.get("currency")
            if not exchange:
                exchange = info.get("exchange", "")
            if name == ticker and info.get("name"):
                name = info["name"]
    except Exception as e:
        logger.warning("Kézi ticker validálás hiba %s: %s", ticker, e)

    purchase_price = data.get("purchase_price")
    purchase_date = str(data.get("purchase_date") or "").strip() or None
    purchase_cost = _non_negative_float_or_none(data.get("purchase_cost"))
    if purchase_cost is None and data.get("purchase_cost") not in (None, ""):
        return jsonify({"ok": False, "error": "Ervenytelen veteli koltseg"}), 400
    purchase_source = str(data.get("purchase_price_source") or "").strip().lower() or None
    price_info = None
    if purchase_price in (None, "") and not cash_currency:
        try:
            price_result = get_prices_for_tickers([ticker])
            price_info = (price_result.get("prices") or {}).get(ticker)
        except Exception as exc:
            logger.warning("Hozzáadási árfolyam fallback hiba %s: %s", ticker, exc)

    if price_info and price_info.get("price") is not None:
        validated = True
        if not currency:
            currency = price_info.get("currency")
        purchase_price = price_info.get("price")
        purchase_source = "current"
    elif purchase_price in (None, "") and info and info.get("last_price"):
        purchase_price = info.get("last_price")
        purchase_source = "current"
    elif purchase_price not in (None, "") and not purchase_source:
        purchase_source = "manual"

    purchase_price_value = _positive_float(purchase_price)
    if purchase_price_value is None:
        return jsonify({"ok": False, "error": "Vételi ár megadása kötelező."}), 400
    purchase_price = purchase_price_value

    if not validated:
        validation_warning = (
            "Az árfolyam most nem elérhető. A részvény hozzáadható, "
            "a vételi árat kézzel is megadhatod."
        )

    uid = current_user_id()
    saved = insert_portfolio_item(uid, {
        "ticker": ticker, "name": name, "qty": qty,
        "currency": currency, "exchange": exchange,
        "source": "manual", "manually_added": True,
        "purchase_price": purchase_price,
        "purchase_date": purchase_date,
        "purchase_cost": purchase_cost,
        "purchase_price_source": purchase_source,
    })

    sym = {
        "ticker": ticker, "name": name, "currency": currency,
        "exchange": exchange, "source": "manual",
        "query_aliases": [ticker.lower(), name.lower()],
        "last_price": (price_info or {}).get("price") or (info.get("last_price") if info else None),
        "last_price_currency": (price_info or {}).get("currency") or currency,
        "last_price_time": (price_info or {}).get("quote_time") or (price_info or {}).get("timestamp"),
    }
    if not cash_currency:
        upsert_symbol_cache(sym)
        symbol_resolver.upsert_symbol(sym)

    log_event(uid, "portfolio_add",
              f"ticker={ticker} qty={qty} manual=true", _client_ip())

    resp = {
        "ok": True, "ticker": ticker, "name": name,
        "currency": currency, "exchange": exchange,
        "validated": validated, "id": saved.get("id"),
        "item": saved,
        "message": "Új vételi tétel hozzáadva.",
    }
    if validation_warning:
        resp["warning"] = validation_warning
    return jsonify(resp)


@app.route("/api/cash", methods=["POST"])
@login_required
def api_cash_add():
    data = request.get_json(silent=True) or {}
    currency = str(data.get("currency") or "").strip().upper()
    if currency not in {"HUF", "EUR", "USD"}:
        return jsonify({"ok": False, "error": "Érvénytelen készpénzdeviza."}), 400
    try:
        amount = float(data.get("amount"))
        if amount <= 0:
            raise ValueError()
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Az összeg legyen nullánál nagyobb."}), 400
    purchase_date = str(data.get("purchase_date") or "").strip() or None
    if purchase_date:
        try:
            datetime.strptime(purchase_date, "%Y-%m-%d")
        except ValueError:
            return jsonify({"ok": False, "error": "A dátum ÉÉÉÉ-HH-NN formátumú legyen."}), 400
    uid = current_user_id()
    ticker = f"CASH-{currency}"
    saved = insert_portfolio_item(uid, {
        "ticker": ticker,
        "name": f"Készpénz ({currency})",
        "qty": amount,
        "currency": currency,
        "exchange": "",
        "source": "cash/manual",
        "manually_added": True,
        "purchase_price": 1,
        "purchase_date": purchase_date,
        "purchase_cost": 0,
        "purchase_price_source": "manual",
    })
    log_event(uid, "portfolio_cash_add", f"ticker={ticker} amount={amount}", _client_ip())
    return jsonify({"ok": True, "item": saved, "message": "Készpénz hozzáadva."})


# ---------------------------------------------------------------------------
# /api/symbols/recent
# ---------------------------------------------------------------------------

@app.route("/api/symbols/recent")
@login_required
def api_symbols_recent():
    db_symbols = db.get_all_symbols(limit=20)
    if db_symbols:
        return jsonify({"symbols": db_symbols})
    symbols = symbol_resolver.get_cached_symbols()
    symbols_sorted = sorted(symbols, key=lambda s: s.get("last_seen", ""), reverse=True)
    return jsonify({"symbols": symbols_sorted[:20]})



# ---------------------------------------------------------------------------
# /api/alerts
# ---------------------------------------------------------------------------

ALERT_LABELS = {
    "price_below": "Részvényárfolyam alá esett",
    "price_above": "Részvényárfolyam elérte / fölé ment",
    "price_change_up_pct": "Részvényárfolyam nőtt",
    "price_change_down_pct": "Részvényárfolyam csökkent",
    "portfolio_value_above": "Portfólióérték elérte / fölé ment",
    "portfolio_value_below": "Portfólióérték alá esett",
    "portfolio_change_up_pct": "Portfólióérték nőtt",
    "portfolio_change_down_pct": "Portfólióérték csökkent",
}


@app.route("/api/alerts", methods=["GET", "POST"])
@login_required
def api_alerts():
    alerts_enabled = get_setting_bool("alerts_enabled", False)
    if request.method == "GET":
        return jsonify({
            "alerts": get_alerts(current_user_id()) if alerts_enabled else [],
            "smtp_configured": smtp_configured(),
            "alerts_enabled": alerts_enabled,
            "default_cooldown_minutes": get_setting_int("alert_cooldown_minutes", 60) or 60,
        })

    if not alerts_enabled:
        return jsonify({"ok": False, "alerts_enabled": False, "error": "Email alerts are disabled"}), 403

    data = request.get_json(silent=True) or {}
    if not data.get("cooldown_minutes"):
        data["cooldown_minutes"] = get_setting_int("alert_cooldown_minutes", 60) or 60
    try:
        alert = create_alert(current_user_id(), data)
        log_event(current_user_id(), "alert_create", f"type={alert.get('alert_type')} ticker={alert.get('ticker') or '-'}", _client_ip())
        return jsonify({"ok": True, "alert": alert, "smtp_configured": smtp_configured()})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@app.route("/api/alerts/<int:alert_id>", methods=["DELETE"])
@login_required
def api_alert_delete(alert_id: int):
    if not get_setting_bool("alerts_enabled", False):
        return jsonify({"ok": False, "alerts_enabled": False, "error": "Email alerts are disabled"}), 403
    ok = delete_alert(current_user_id(), alert_id)
    if ok:
        log_event(current_user_id(), "alert_delete", f"id={alert_id}", _client_ip())
    return jsonify({"ok": ok})


@app.route("/api/alerts/<int:alert_id>/toggle", methods=["POST"])
@login_required
def api_alert_toggle(alert_id: int):
    if not get_setting_bool("alerts_enabled", False):
        return jsonify({"ok": False, "alerts_enabled": False, "error": "Email alerts are disabled"}), 403
    data = request.get_json(silent=True) or {}
    active = bool(data.get("active"))
    ok = set_alert_active(current_user_id(), alert_id, active)
    if ok:
        log_event(current_user_id(), "alert_toggle", f"id={alert_id} active={active}", _client_ip())
    return jsonify({"ok": ok})


@app.route("/api/alerts/check", methods=["POST"])
@login_required
def api_alert_check():
    if not get_setting_bool("alerts_enabled", False):
        return jsonify({"ok": True, "alerts_enabled": False, "message": "Email alerts are disabled"})

    uid = current_user_id()
    alerts = get_alerts(uid, active_only=True)
    if not alerts:
        return jsonify({"ok": True, "checked": 0, "triggered": [], "warnings": []})

    portfolio = get_portfolio(uid)
    tickers = [i["ticker"] for i in portfolio]
    prices_result = get_prices_for_tickers(tickers) if tickers else {"prices": {}}
    prices = prices_result.get("prices", {})
    fx_result = get_fx_rates(get_setting("fx_rate_mode", "market"))
    fx = fx_result.get("fx", {})

    triggered, warnings = _evaluate_alerts(uid, alerts, portfolio, prices, fx)
    return jsonify({
        "ok": True,
        "alerts_enabled": True,
        "checked": len(alerts),
        "triggered": triggered,
        "warnings": warnings,
        "smtp_configured": smtp_configured(),
    })


def _evaluate_alerts(uid: int, alerts: list[dict], portfolio: list[dict], prices: dict, fx: dict):
    triggered = []
    warnings = []
    portfolio_total_huf = _portfolio_total_huf(portfolio, prices, fx)

    for alert in alerts:
        current_value = _alert_current_value(alert, portfolio_total_huf, prices, fx)
        if current_value is None:
            update_alert_state(uid, alert["id"], checked=True)
            continue

        alert_type = alert.get("alert_type")
        should_trigger = False
        threshold = alert.get("threshold")
        percent = alert.get("percent")
        last_value = alert.get("last_value")

        if alert_type in ("price_above", "portfolio_value_above"):
            should_trigger = threshold is not None and current_value >= float(threshold)
        elif alert_type in ("price_below", "portfolio_value_below"):
            should_trigger = threshold is not None and current_value <= float(threshold)
        elif alert_type in ("price_change_up_pct", "portfolio_change_up_pct"):
            if last_value is None:
                update_alert_state(uid, alert["id"], last_value=current_value, checked=True)
                continue
            should_trigger = current_value >= float(last_value) * (1 + float(percent) / 100.0)
        elif alert_type in ("price_change_down_pct", "portfolio_change_down_pct"):
            if last_value is None:
                update_alert_state(uid, alert["id"], last_value=current_value, checked=True)
                continue
            should_trigger = current_value <= float(last_value) * (1 - float(percent) / 100.0)

        if should_trigger and _alert_can_trigger(alert):
            subject, body = _alert_email_text(alert, current_value)
            ok, msg = send_email(alert.get("email_to"), subject, body)
            if ok:
                update_alert_state(uid, alert["id"], last_value=current_value, triggered=True, checked=True)
                triggered.append({
                    "id": alert["id"],
                    "label": ALERT_LABELS.get(alert_type, alert_type),
                    "ticker": alert.get("ticker"),
                    "current_value": current_value,
                    "currency": _alert_display_currency(alert),
                })
                log_event(uid, "alert_triggered", f"id={alert['id']} type={alert_type} value={current_value}", _client_ip())
            else:
                warnings.append(msg)
                update_alert_state(uid, alert["id"], checked=True)
        else:
            # Százalékos riasztásnál a bázisértéket megtartjuk az első ellenőrzéstől
            # vagy az utolsó sikeres riasztástól számítva. Így nem nullázódik minden frissítéskor.
            if "change" in alert_type:
                update_alert_state(uid, alert["id"], checked=True)
            else:
                update_alert_state(uid, alert["id"], last_value=current_value, checked=True)

    return triggered, warnings


def _alert_current_value(alert: dict, portfolio_total_huf: float | None, prices: dict, fx: dict):
    t = alert.get("alert_type") or ""
    if t.startswith("portfolio_"):
        return portfolio_total_huf

    ticker = (alert.get("ticker") or "").strip().upper()
    if not ticker or ticker not in prices:
        return None
    p = prices.get(ticker) or {}
    if p.get("price") is None:
        return None
    from_currency = (p.get("currency") or alert.get("currency") or "").upper()
    to_currency = (alert.get("currency") or from_currency or "HUF").upper()
    return _convert_currency(float(p["price"]), from_currency, to_currency, fx)


def _portfolio_total_huf(portfolio: list[dict], prices: dict, fx: dict):
    total = 0.0
    has_data = False
    for item in portfolio:
        p = prices.get(item.get("ticker")) or {}
        if p.get("price") is None:
            continue
        qty = float(item.get("qty") or 0)
        currency = (p.get("currency") or item.get("currency") or "").upper()
        huf = _convert_currency(float(p["price"]) * qty, currency, "HUF", fx)
        if huf is not None:
            total += huf
            has_data = True
    return total if has_data else None


def _convert_currency(value: float, from_currency: str, to_currency: str, fx: dict):
    from_currency = (from_currency or "").upper()
    to_currency = (to_currency or "").upper()
    if not from_currency or not to_currency:
        return value
    if from_currency == to_currency:
        return value
    if from_currency == "GBX":
        value = value / 100.0
        from_currency = "GBP"
    if to_currency == "GBX":
        gbp = _convert_currency(value, from_currency, "GBP", fx)
        return gbp * 100.0 if gbp is not None else None
    if to_currency == "HUF":
        if from_currency == "HUF":
            return value
        rate = fx.get(f"{from_currency}/HUF")
        return value * rate if rate else None
    if from_currency == "HUF":
        rate = fx.get(f"{to_currency}/HUF")
        return value / rate if rate else None
    huf = _convert_currency(value, from_currency, "HUF", fx)
    if huf is None:
        return None
    return _convert_currency(huf, "HUF", to_currency, fx)


def _alert_can_trigger(alert: dict) -> bool:
    last = alert.get("last_triggered_at")
    if not last:
        return True
    try:
        last_dt = datetime.fromisoformat(str(last).replace("Z", "+00:00")).replace(tzinfo=None)
    except Exception:
        return True
    cooldown = int(alert.get("cooldown_minutes") or 60)
    return datetime.now() - last_dt >= timedelta(minutes=cooldown)


def _alert_display_currency(alert: dict) -> str:
    if (alert.get("alert_type") or "").startswith("portfolio_"):
        return "HUF"
    return (alert.get("currency") or "").upper() or ""


def _alert_email_text(alert: dict, current_value: float):
    label = ALERT_LABELS.get(alert.get("alert_type"), alert.get("alert_type"))
    currency = _alert_display_currency(alert)
    ticker = alert.get("ticker") or "Teljes portfólió"
    value_txt = f"{current_value:,.2f} {currency}".replace(",", " ")

    parts = [
        "Portfólió Követő riasztás",
        "",
        f"Riasztás: {label}",
        f"Érintett: {ticker}",
        f"Aktuális érték: {value_txt}",
    ]
    if alert.get("threshold") is not None:
        parts.append(f"Beállított célérték: {float(alert['threshold']):,.2f} {currency}".replace(",", " "))
    if alert.get("percent") is not None:
        parts.append(f"Beállított változás: {float(alert['percent']):.2f}%")
    if alert.get("last_value") is not None:
        parts.append(f"Bázisérték: {float(alert['last_value']):,.2f} {currency}".replace(",", " "))
    parts.extend([
        "",
        "Megjegyzés: ez automatikus értesítés. A piaci adatok külső szolgáltatóktól érkeznek, ezért döntés előtt érdemes ellenőrizni őket.",
    ])
    return f"Portfólió riasztás: {label}", "\n".join(parts)

# ---------------------------------------------------------------------------
# /api/export/xlsx
# ---------------------------------------------------------------------------

@app.route("/api/export/xlsx")
@login_required
def api_export_xlsx():
    if not get_bool("excel_export_enabled"):
        return jsonify({"error": "Excel export ki van kapcsolva."}), 403

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl nincs telepítve."}), 500

    uid = current_user_id()
    portfolio = get_portfolio(uid)
    if not portfolio:
        return jsonify({"error": "Üres portfólió"}), 400

    tickers = [i["ticker"] for i in portfolio]
    price_result = get_prices_for_tickers(tickers)
    prices = price_result.get("prices", {})

    fx_result = get_fx_rates(get_setting("fx_rate_mode", "market"))
    fx = fx_result.get("fx", {})

    def to_huf(val, currency):
        if val is None:
            return None
        c = (currency or "").upper()
        if c == "HUF":
            return round(val, 2)
        if c in ("GBP", "GBX") and fx.get("GBP/HUF"):
            return round(val * fx["GBP/HUF"] / (100 if c == "GBX" else 1), 2)
        rate = fx.get(f"{c}/HUF")
        return round(val * rate, 2) if rate else None

    wb = Workbook()
    ws = wb.active
    ws.title = "Portfólió"

    headers = [
        "Részvény neve", "Ticker", "Tőzsde", "Darab",
        "Aktuális árfolyam", "Deviza", "Forrás",
        "Érték (saját deviza)", "Érték (HUF)", "Érték (EUR)", "Érték (USD)",
        "Árfolyam időpontja", "Elavult ár?", "Hozzáadás módja", "Export időpontja",
    ]
    headers[4:4] = ["Vétel dátuma", "Vételi ár", "Vételi költség"]
    headers[14:14] = ["Befektetett érték", "Nyereség / veszteség", "Hozam %"]

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1A1A2E")
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    export_ts = datetime.now().strftime("%Y.%m.%d %H:%M")
    total_huf = 0.0

    for item in portfolio:
        ticker = item["ticker"]
        p = prices.get(ticker)
        price = p["price"] if p else (item.get("last_price"))
        currency = (p["currency"] if p else None) or item.get("currency") or ""
        val_own = round(price * item["qty"], 4) if price is not None else None
        val_huf = to_huf(val_own, currency)
        val_eur = round(val_huf / fx["EUR/HUF"], 2) if val_huf and fx.get("EUR/HUF") else None
        val_usd = round(val_huf / fx["USD/HUF"], 2) if val_huf and fx.get("USD/HUF") else None
        purchase_price = item.get("purchase_price")
        purchase_cost = item.get("purchase_cost")
        try:
            purchase_price = float(purchase_price) if purchase_price not in (None, "") else None
        except (TypeError, ValueError):
            purchase_price = None
        try:
            purchase_cost = float(purchase_cost) if purchase_cost not in (None, "") else 0.0
        except (TypeError, ValueError):
            purchase_cost = 0.0
        purchase_cost = max(purchase_cost, 0.0)
        invested = None
        profit_loss = None
        return_pct = None
        if purchase_price is not None and purchase_price > 0:
            invested = round(purchase_price * item["qty"] + purchase_cost, 4)
            if val_own is not None:
                profit_loss = round(val_own - invested, 4)
                return_pct = round((profit_loss / invested) * 100, 2) if invested > 0 else None
        if val_huf:
            total_huf += val_huf
        ws.append([
            item.get("name", ticker), ticker, item.get("exchange", ""), item["qty"],
            item.get("purchase_date", ""), purchase_price, purchase_cost if purchase_cost else None,
            price, currency, (p or {}).get("source", "cache" if item.get("last_price") else ""),
            val_own, val_huf, val_eur, val_usd,
            invested, profit_loss, return_pct,
            (p or {}).get("quote_time") or (p or {}).get("timestamp") or item.get("last_price_time", ""),
            "Igen" if (p or {}).get("stale") else "Nem",
            "kézi" if item.get("manually_added") else "keresés",
            export_ts,
        ])

    ws.append([])
    ws.append(["ÖSSZESÍTÉS"])
    ws.append(["Összesen HUF", total_huf])
    if fx.get("EUR/HUF") and total_huf:
        ws.append(["Összesen EUR", round(total_huf / fx["EUR/HUF"], 2)])
    if fx.get("USD/HUF") and total_huf:
        ws.append(["Összesen USD", round(total_huf / fx["USD/HUF"], 2)])
    ws.append(["Árfolyam nélküli tételek", sum(1 for i in portfolio if i["ticker"] not in prices)])
    market_fx = fx_result.get("market") or {}
    official_fx = fx_result.get("official") or {}
    ws.append(["Hasznalt devizaarfolyam mod", fx_result.get("mode", "")])
    ws.append(["Kert devizaarfolyam mod", fx_result.get("requested_mode", fx_result.get("mode", ""))])
    ws.append(["EUR/HUF hasznalt arfolyam", fx.get("EUR/HUF")])
    ws.append(["USD/HUF hasznalt arfolyam", fx.get("USD/HUF")])
    ws.append(["MNB EUR/HUF", official_fx.get("EUR/HUF")])
    ws.append(["MNB USD/HUF", official_fx.get("USD/HUF")])
    ws.append(["Piaci EUR/HUF", market_fx.get("EUR/HUF")])
    ws.append(["Piaci USD/HUF", market_fx.get("USD/HUF")])
    ws.append(["Forrás (deviza)", fx_result.get("source", "")])
    ws.append(["Devizaarfolyam idopontja", fx_result.get("timestamp", "")])

    col_widths = [22, 12, 12, 8, 16, 8, 18, 18, 18, 14, 14, 20, 10, 14, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "#,##0.00"
            cell.alignment = Alignment(vertical="center")
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    log_event(current_user_id(), "excel_export", f"sorok={len(portfolio)}", _client_ip())

    filename = datetime.now().strftime("portfolio_%Y%m%d_%H%M.xlsx")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ===========================================================================
# ADMIN FELÜLET
# ===========================================================================

@app.route("/admin")
@admin_required
def admin_index():
    stats = get_stats()
    return render_template("admin/index.html",
                           stats=stats, username=current_username(), role=current_role())


# ---------------------------------------------------------------------------
# /admin/users
# ---------------------------------------------------------------------------

@app.route("/admin/users")
@admin_required
def admin_users():
    users = get_all_users()
    return render_template("admin/users.html", users=users,
                           username=current_username(), role=current_role())


@app.route("/admin/users/new", methods=["GET", "POST"])
@admin_required
def admin_user_new():
    error = None
    if request.method == "POST":
        uname = request.form.get("username", "").strip()
        pw1 = request.form.get("password", "")
        pw2 = request.form.get("password2", "")
        role = request.form.get("role", "user")
        is_active = request.form.get("is_active", "1") == "1"

        if not uname or not pw1:
            error = "Felhasználónév és jelszó kötelező."
        elif pw1 != pw2:
            error = "A két jelszó nem egyezik."
        elif len(pw1) < 6:
            error = "A jelszó legalább 6 karakter legyen."
        else:
            try:
                create_user(uname, pw1, role=role, is_active=is_active)
                log_event(current_user_id(), "user_create",
                          f"username={uname} role={role}", _client_ip())
                flash(f"Felhasználó létrehozva: {uname}", "success")
                return redirect(url_for("admin_users"))
            except Exception as e:
                error = f"Hiba: {e}"

    return render_template("admin/user_form.html", error=error, edit_user=None,
                           username=current_username(), role=current_role())


@app.route("/admin/users/<int:uid>/edit", methods=["GET", "POST"])
@admin_required
def admin_user_edit(uid: int):
    edit_user = get_user_by_id(uid)
    if not edit_user:
        flash("Felhasználó nem található.", "error")
        return redirect(url_for("admin_users"))

    error = None
    if request.method == "POST":
        action = request.form.get("action", "")
        if action == "update":
            new_role = request.form.get("role", edit_user["role"])
            new_active = 1 if request.form.get("is_active") == "1" else 0
            update_user(uid, role=new_role, is_active=new_active)
            log_event(current_user_id(), "user_update",
                      f"uid={uid} role={new_role} active={new_active}", _client_ip())
            flash("Felhasználó módosítva.", "success")
            return redirect(url_for("admin_users"))
        elif action == "password":
            pw1 = request.form.get("password", "")
            pw2 = request.form.get("password2", "")
            if pw1 != pw2:
                error = "A két jelszó nem egyezik."
            elif len(pw1) < 6:
                error = "A jelszó legalább 6 karakter legyen."
            else:
                set_user_password(uid, pw1)
                log_event(current_user_id(), "user_password_change",
                          f"uid={uid}", _client_ip())
                flash("Jelszó módosítva.", "success")
                return redirect(url_for("admin_users"))

    edit_user = get_user_by_id(uid)
    return render_template("admin/user_form.html", error=error, edit_user=edit_user,
                           username=current_username(), role=current_role())


# ---------------------------------------------------------------------------
# /admin/settings
# ---------------------------------------------------------------------------

@app.route("/admin/settings", methods=["GET", "POST"])
@admin_required
def admin_settings():
    if request.method == "POST":
        for key in request.form:
            if key.startswith("setting_"):
                setting_key = key[len("setting_"):]
                value = request.form[key].strip()
                if setting_key == "auto_refresh_seconds" and value not in {"0", "60", "300", "900", "1800", "3600"}:
                    value = "300"
                save_setting(setting_key, value)
        if "setting_alerts_enabled" not in request.form:
            save_setting("alerts_enabled", "false")
        log_event(current_user_id(), "settings_update", "", _client_ip())
        flash("Beállítások mentve.", "success")
        return redirect(url_for("admin_settings"))

    settings = get_all_settings_with_defaults()
    settings_map = {s["key"]: s for s in settings}
    return render_template("admin/settings.html", settings=settings, settings_map=settings_map,
                           username=current_username(), role=current_role())


# ---------------------------------------------------------------------------
# /admin/system
# ---------------------------------------------------------------------------

@app.route("/admin/system")
@admin_required
def admin_system():
    import platform
    from services.db import _is_postgres
    from services.fx import FX_FILE_CACHE

    stats = get_stats()

    # DB kapcsolat check
    db_ok = True
    try:
        db.get_stats()
    except Exception:
        db_ok = False

    # FX utolsó frissítés
    fx_cached_ts = None
    try:
        if os.path.exists(FX_FILE_CACHE):
            import json as _json
            with open(FX_FILE_CACHE, "r", encoding="utf-8") as f:
                fxd = _json.load(f)
                fx_cached_ts = fxd.get("timestamp")
    except Exception:
        pass

    def is_set(key):
        return bool(os.environ.get(key, "").strip())

    status = {
        "app_version": "1.0.0",
        "python_version": platform.python_version(),
        "db_type": "PostgreSQL" if _is_postgres() else "SQLite",
        "db_ok": db_ok,
        "app_env": os.environ.get("APP_ENV", "local"),
        "has_secret_key": is_set("SECRET_KEY"),
        "has_database_url": is_set("DATABASE_URL"),
        "has_portfolio_password": is_set("PORTFOLIO_PASSWORD"),
        "default_admin_danger": (
            os.environ.get("PORTFOLIO_USERNAME", "admin") == "admin"
            and os.environ.get("PORTFOLIO_PASSWORD", "admin") == "admin"
        ),
        "fx_last_update": fx_cached_ts,
        "enable_yahoo": get_bool("enable_yahoo"),
        "enable_stooq": get_bool("enable_stooq"),
        "enable_mnb": get_bool("enable_mnb"),
        "maintenance_mode": get_bool("maintenance_mode"),
        **stats,
    }
    return render_template("admin/system.html", status=status,
                           username=current_username(), role=current_role())


# ---------------------------------------------------------------------------
# /admin/logs
# ---------------------------------------------------------------------------

@app.route("/admin/logs")
@admin_required
def admin_logs():
    logs = get_audit_logs(limit=100)
    return render_template("admin/logs.html", logs=logs,
                           username=current_username(), role=current_role())


# ---------------------------------------------------------------------------
# Hibaoldalak
# ---------------------------------------------------------------------------

@app.errorhandler(403)
def forbidden(_e):
    return render_template("403.html"), 403


@app.errorhandler(404)
def not_found(_e):
    return render_template("404.html") if os.path.exists("templates/404.html") else ("Oldal nem található.", 404)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _ts() -> str:
    return datetime.now().isoformat(timespec="seconds")


# ---------------------------------------------------------------------------
# Indítás
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    init_db()
    init_default_settings()

    use_reloader = os.environ.get("FLASK_USE_RELOADER", "false").lower() == "true"

    print("=" * 62)
    print("  Portfólió Követő indul...")
    print("  Nyisd meg: http://localhost:5000")
    print("  Belépés: admin / admin (fejlesztés) vagy .env")
    print(f"  Flask reloader: {'BE' if use_reloader else 'KI'}")
    print(f"  DB: {'PostgreSQL' if db._is_postgres() else 'SQLite'}")
    print("=" * 62)

    app.run(
        host="127.0.0.1",
        port=5000,
        debug=True,
        use_reloader=use_reloader,
    )
